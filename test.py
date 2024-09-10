import torch
import os
from envs import VFEnv

import utils

from config import Configuration
from memoryreplay import MemoryReplay
from trainer import Trainer
from models import RLAgent
import openpyxl

if __name__ == '__main__':

    # Get configuration
    config = Configuration("config.json")

    fn = config.output_path + config.suffix + ".xlsx"
    sheet_name = "stop-std="+str(config.stop_std)
    if not os.path.exists(fn):
        wb = openpyxl.Workbook()
    else:
        wb = openpyxl.load_workbook(fn)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        wb.remove_sheet(ws)    
    ws = wb.create_sheet(sheet_name)
    vals = []
    column_names = ["seed", "steps", "mse"]
    ws.append(column_names)
    pred_path = './runs/'
    for seed in [100, 2023, 20000, 50000, 100000]:
        utils.fix_seed(seed)
        env = VFEnv(init_fos_std=config.init_fos_std, config=config)

        # Global initialization
        torch.cuda.init()
        device = torch.device(config.device if torch.cuda.is_available() else "cpu")

        # Information about environments
        state_dim = env.state_dim
        action_dim = env.action_dim

        # Prepare Experience Memory Replay
        memory = MemoryReplay(config.capacity)

        suffix = config.suffix
        path = './runs/{}/{}/'.format(config.type+"_"+suffix+"_zest", config.data_name)

        
        model = RLAgent(
            state_dim=state_dim,
            action_dim=action_dim,
            target_update_freq=config.target_update_freq,
            learning_rate=config.lr,
            gamma=config.gamma,
            hidden_dim=config.hidden_dim,
            td_target=config.td_target,
            device=device
        )
        utils.load_model(model, path, str(seed) + "_std" + str(config.stop_std), suffix, "best")
        model.eval()

        trainer = Trainer(
            model=model,
            env=env,
            memory=memory,
            epsilon_start=config.epsilon_start,
            epsilon_final=config.epsilon_final,
            epsilon_decay=config.epsilon_decay,
            start_learning=config.start_learning,
            batch_size=config.batch_size,
            save_update_freq=config.save_update_freq,
            output_dir=path,
            config=config,
            phase='test',
            output_path=config.output_path,
            suffix=config.suffix + '_' + str(seed) + "_std" + str(config.stop_std)
        )
        random = False # True for random location and initial stimulus value
        mse, steps = trainer.test("test", rand=random)
        vals.append([seed, steps, mse])
    for row in vals:
        ws.append(row)
    wb.save(fn)
    
